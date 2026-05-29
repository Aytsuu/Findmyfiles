from __future__ import annotations

from pathlib import Path

try:
    from docx import Document
except ImportError:  # pragma: no cover
    Document = None

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


def extract_docx_text(path: str | Path) -> str:
    if Document is None:
        msg = "python-docx is required for DOCX extraction"
        raise RuntimeError(msg)
    document = Document(path)
    return "\n".join(paragraph.text for paragraph in document.paragraphs if paragraph.text).strip()


def extract_xlsx_text(path: str | Path) -> str:
    if load_workbook is None:
        msg = "openpyxl is required for XLSX extraction"
        raise RuntimeError(msg)
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            values = [str(value) for value in row if value is not None]
            if values:
                rows.append(" ".join(values))
    return "\n".join(rows).strip()
